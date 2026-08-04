"""Strict local reader for the SERDP09 narrowband far-field archive."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import math
import re
from typing import Literal

import numpy as np
from openpyxl import load_workbook

from .physics import EngineState, JetSource, JetStream, THIRD_OCTAVE_HZ


README_NAME = "0SERDP09acousticREADME.txt"
WORKBOOK_NAME = "SERDP09acousticEscort.xlsx"
SHEET_NAME = "SERDP09_FF.db"
_VARIABLES = ('"Hz"', '"Polar angle"', '"PSD (dB)"')
_ZONE_RE = re.compile(r'^ZONE\s+T="(?P<title>[^"\r\n]+)"\s+I=(?P<i>\d+)\s+J=(?P<j>\d+)\s+F=POINT\s*$', re.I)
_AUXDATA_RE = re.compile(r"^(?:DATASET)?AUXDATA\b", re.I)


class Serdp09ReferenceError(ValueError):
    """Raised when a local SERDP09 archive cannot support a comparison."""


@dataclass(frozen=True, slots=True)
class Serdp09Record:
    """One explicitly selected workbook record."""

    icrdg: int
    metadata: tuple[tuple[str, str], ...]


@dataclass(frozen=True, slots=True)
class Serdp09Case:
    """Third-octave levels derived from one narrowband directivity file."""

    record: Serdp09Record
    angles_deg: tuple[float, ...]
    bands_db: dict[float, np.ndarray]


@dataclass(frozen=True, slots=True)
class Serdp09ShapeMetric:
    """Peak-normalized spectral-shape residual at one measurement angle."""

    angle_deg: float
    rmse_db: float


@dataclass(frozen=True, slots=True)
class Serdp09BroadbandLevel:
    """Once-normalized broadband level pair at one measurement angle."""

    angle_deg: float
    observed_db: float
    predicted_db: float


@dataclass(frozen=True, slots=True)
class Serdp09Evaluation:
    """Auditable fixed-physics SERDP09 comparison result."""

    verdict: Literal["compatible", "incompatible"]
    reasons: tuple[str, ...]
    fixed_parameters: tuple[tuple[str, float], ...]
    reference_angle_deg: float | None
    normalization_order: tuple[str, ...]
    shape_rmse_db_by_angle: tuple[Serdp09ShapeMetric, ...]
    broadband_levels_by_angle: tuple[Serdp09BroadbandLevel, ...]
    broadband_directivity_rmse_db: float | None
    exclusions: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class Serdp09TrustedProof:
    """Trusted synthetic evidence supplied by an evaluator caller, never archive data."""

    stream_mapping_proven: bool
    frequency_basis_proven: bool


def _fail(detail: str) -> None:
    raise Serdp09ReferenceError(detail)


def _case_id(value: int | str) -> int:
    """Accept only a decimal archive identifier, never a filesystem fragment."""
    if isinstance(value, bool):
        _fail("icrdg must be an integer")
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.isdecimal():
        return int(value)
    _fail("icrdg must be an integer")


def _required_file(root: Path, name: str) -> Path:
    path = root / name
    if not path.is_file():
        _fail(f"missing required archive file: {name}")
    return path


def _load_record(root: Path, icrdg: int) -> Serdp09Record:
    readme = _required_file(root, README_NAME)
    try:
        readme_text = readme.read_text(encoding="utf-8")
    except UnicodeDecodeError as error:
        raise Serdp09ReferenceError("archive README is not UTF-8") from error
    if "icrdg" not in readme_text.lower():
        _fail("archive README does not define icrdg")
    workbook_path = _required_file(root, WORKBOOK_NAME)
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            if SHEET_NAME not in workbook.sheetnames:
                _fail(f"workbook missing sheet: {SHEET_NAME}")
            sheet = workbook[SHEET_NAME]
            values = sheet.iter_rows(values_only=True)
            header = next(values, None)
            if header is None:
                _fail("workbook has no header row")
            columns = tuple(str(value).strip() for value in header)
            if "icrdg" not in columns:
                _fail("workbook missing required icrdg column")
            selected = [row for row in values if row[columns.index("icrdg")] == icrdg]
        finally:
            workbook.close()
    except OSError as error:
        raise Serdp09ReferenceError("unable to read workbook") from error
    if not selected:
        _fail(f"icrdg {icrdg} not found in workbook")
    if len(selected) != 1:
        _fail(f"duplicate workbook record for icrdg {icrdg}")
    return Serdp09Record(
        icrdg=icrdg,
        metadata=tuple((column, str(value)) for column, value in zip(columns, selected[0])),
    )


def _parse_zone(line: str, icrdg: int) -> tuple[int, int]:
    match = _ZONE_RE.fullmatch(line)
    if match is None:
        _fail("invalid ZONE header")
    if match.group("title") != str(icrdg):
        _fail("ZONE identifier does not match icrdg")
    count, angle_count = int(match.group("i")), int(match.group("j"))
    if count <= 0 or angle_count <= 0:
        _fail("ZONE dimensions must be positive")
    return count, angle_count


def _parse_spectrum(path: Path, icrdg: int) -> dict[float, tuple[np.ndarray, np.ndarray]]:
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except UnicodeDecodeError as error:
        raise Serdp09ReferenceError("spectrum file is not UTF-8") from error
    variables_index = next(
        (index for index, line in enumerate(lines) if line.upper().startswith("VARIABLES")),
        None,
    )
    if variables_index is None or not all(token in lines[variables_index] for token in _VARIABLES):
        _fail("spectrum columns must be Hz, Polar angle, PSD (dB)")
    zone_index = next(
        (index for index in range(variables_index + 1, len(lines)) if lines[index].upper().startswith("ZONE")),
        None,
    )
    if zone_index is None:
        _fail("invalid ZONE header")
    count, angle_count = _parse_zone(lines[zone_index], icrdg)
    data = [
        line.split() for line in lines[zone_index + 1:]
        if line.strip() and _AUXDATA_RE.match(line) is None
    ]
    if len(data) != count * angle_count:
        _fail("spectrum row count does not match ZONE dimensions")
    triples: list[tuple[float, float, float]] = []
    for fields in data:
        if len(fields) != 3:
            _fail("spectrum row must have exactly three columns")
        try:
            frequency, angle, psd = (float(field) for field in fields)
        except ValueError as error:
            raise Serdp09ReferenceError("spectrum contains non-numeric data") from error
        if -999.0 in (frequency, angle, psd):
            _fail("spectrum contains -999 invalid data")
        if not all(math.isfinite(value) for value in (frequency, angle, psd)):
            _fail("spectrum contains non-finite data")
        triples.append((frequency, angle, psd))
    if len(set(triples)) != len(triples):
        _fail("spectrum contains duplicate rows")
    groups: dict[float, tuple[np.ndarray, np.ndarray]] = {}
    for start in range(0, len(triples), count):
        group = triples[start:start + count]
        angle = group[0][1]
        if any(row[1] != angle for row in group):
            _fail("spectrum angle grouping does not match ZONE dimensions")
        if angle in groups:
            _fail("spectrum contains duplicate angles")
        frequencies = np.array([row[0] for row in group], dtype=float)
        psd = np.array([row[2] for row in group], dtype=float)
        if not np.all(np.diff(frequencies) > 0.0):
            _fail("spectrum frequency must increase within each angle")
        groups[angle] = (frequencies, psd)
    if len(groups) != angle_count:
        _fail("spectrum angle count does not match ZONE dimensions")
    return groups


def _integrate_third_octaves(
    frequencies: np.ndarray, psd_db: np.ndarray
) -> np.ndarray:
    lower = THIRD_OCTAVE_HZ / 2.0 ** (1.0 / 6.0)
    upper = THIRD_OCTAVE_HZ * 2.0 ** (1.0 / 6.0)
    if lower[0] < frequencies[0] or upper[-1] > frequencies[-1]:
        _fail("spectrum has no overlap with all third-octave bands")
    linear = 10.0 ** (psd_db / 10.0)
    bands: list[float] = []
    for lo, hi in zip(lower, upper):
        interior = (frequencies > lo) & (frequencies < hi)
        grid = np.concatenate(([lo], frequencies[interior], [hi]))
        values = np.interp(grid, frequencies, linear)
        energy = np.trapezoid(values, grid)
        if not math.isfinite(float(energy)) or energy <= 0.0:
            _fail("third-octave integration produced invalid energy")
        bands.append(10.0 * math.log10(float(energy)))
    return np.array(bands, dtype=float)


def load_case(data_root: Path, icrdg: int | str) -> Serdp09Case:
    """Load one explicit local archive case and integrate its PSD by angle."""
    root = Path(data_root)
    if not root.is_dir():
        _fail("archive root does not exist")
    case_id = _case_id(icrdg)
    record = _load_record(root, case_id)
    filename = root / f"{case_id}_nb" / f"{case_id}_nb.dat"
    if not filename.is_file():
        _fail(f"missing spectrum file for icrdg {case_id}")
    spectra = _parse_spectrum(filename, case_id)
    bands = {angle: _integrate_third_octaves(*spectrum) for angle, spectrum in spectra.items()}
    return Serdp09Case(record, tuple(spectra), bands)


_FIXED_JET_PARAMETERS = (
    ("c_jet_db", 140.0),
    ("frequency_scale", 1.0),
    ("outer_velocity_ms", 360.0),
    ("outer_mass_flow_kg_s", 250.0),
    ("outer_diameter_m", 1.5),
    ("outer_temperature_k", 450.0),
    ("outer_total_pressure_pa", 175000.0),
    ("merged_velocity_ms", 330.0),
    ("merged_mass_flow_kg_s", 320.0),
    ("merged_diameter_m", 1.7),
    ("merged_temperature_k", 430.0),
    ("merged_total_pressure_pa", 175000.0),
)
_NORMALIZATION_ORDER = (
    "per-angle spectral peak normalization",
    "broadband energetic sum",
    "single 90.0 degree angular normalization",
)
_EXCLUSIONS = (
    "absolute level comparison excluded",
    "calibration and parameter fitting excluded",
    "simplified mixed-jet fallback excluded",
)


def _compatibility_reasons(trusted_proof: Serdp09TrustedProof | None) -> tuple[str, ...]:
    reasons: list[str] = []
    if trusted_proof is None or not trusted_proof.stream_mapping_proven:
        reasons.append("missing proven core/bypass-to-outer/merged mapping")
    if trusted_proof is None or not trusted_proof.frequency_basis_proven:
        reasons.append("missing proven model-scale frequency basis")
    return tuple(reasons)


def _fixed_engine_state() -> EngineState:
    """Create the declared detailed-stream state without a mixed-jet path."""
    return EngineState(
        1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0,
        jet_streams=(
            JetStream(360.0, 250.0, 1.5, 450.0, 175000.0, "outer"),
            JetStream(330.0, 320.0, 1.7, 430.0, 175000.0, "merged"),
        ),
    )


def _energetic_sum(levels_db: np.ndarray) -> float:
    return float(10.0 * np.log10(np.sum(10.0 ** (levels_db / 10.0))))


def evaluate_case(
    case: Serdp09Case,
    *,
    trusted_proof: Serdp09TrustedProof | None = None,
) -> Serdp09Evaluation:
    """Evaluate a case; the CLI never supplies trusted proof evidence."""
    reasons = _compatibility_reasons(trusted_proof)
    if reasons:
        return Serdp09Evaluation(
            "incompatible", reasons, _FIXED_JET_PARAMETERS, None,
            _NORMALIZATION_ORDER, (), (), None, _EXCLUSIONS,
        )
    state = _fixed_engine_state()
    source = JetSource(c_jet=140.0, f_scale=1.0)
    predicted: dict[float, np.ndarray] = {}
    angles = tuple(sorted(case.angles_deg))
    for angle in angles:
        pieces, status = source.component_spectra_with_diagnostics(state, angle)
        if not pieces or not status.complete:
            _fail("fixed detailed jet source is unavailable")
        stack = np.stack(tuple(pieces.values()))
        predicted[angle] = 10.0 * np.log10(np.sum(10.0 ** (stack / 10.0), axis=0))
    reference_angle = 90.0 if 90.0 in case.bands_db else angles[0]
    shape: list[Serdp09ShapeMetric] = []
    broadband: list[Serdp09BroadbandLevel] = []
    observed_broadband: dict[float, float] = {}
    predicted_broadband: dict[float, float] = {}
    for angle in angles:
        observed = case.bands_db[angle]
        model = predicted[angle]
        if observed.shape != THIRD_OCTAVE_HZ.shape:
            _fail("evaluation spectrum does not match third-octave basis")
        if not np.all(np.isfinite(observed)) or not np.all(np.isfinite(model)):
            _fail("evaluation contains non-finite spectrum values")
        residual = (observed - float(np.max(observed))) - (model - float(np.max(model)))
        shape.append(Serdp09ShapeMetric(angle, float(np.sqrt(np.mean(residual ** 2)))))
        observed_broadband[angle] = _energetic_sum(observed)
        predicted_broadband[angle] = _energetic_sum(model)
    observed_reference = observed_broadband[reference_angle]
    predicted_reference = predicted_broadband[reference_angle]
    for angle in angles:
        broadband.append(Serdp09BroadbandLevel(
            angle, observed_broadband[angle] - observed_reference,
            predicted_broadband[angle] - predicted_reference,
        ))
    directivity_residual = np.array(
        [level.observed_db - level.predicted_db for level in broadband], dtype=float)
    return Serdp09Evaluation(
        "compatible", (), _FIXED_JET_PARAMETERS, reference_angle,
        _NORMALIZATION_ORDER, tuple(shape), tuple(broadband),
        float(np.sqrt(np.mean(directivity_residual ** 2))), _EXCLUSIONS,
    )
